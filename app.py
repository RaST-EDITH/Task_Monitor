# Library and Modules used 
import os                                                    # pip install os ( In case if not present )
import numpy as np                                           # pip install numpy ( In case if not present )
import pandas as pd                                          # pip install pandas==1.4.3
import openpyxl as oxl                                       # pip install openpyxl==3.0.10
from tkinter import *                                        # pip install tkinter==8.6
import customtkinter as ctk                                  # pip install customtkinter==4.6.3
from PIL import Image ,ImageTk                               # pip install pillow==9.3.0
import matplotlib.pyplot as plt                              # pip install matplotlib==3.4.3
from tkinter.messagebox import showerror, showinfo

class TaskMonitor :

    def __init__(self) :
        ctk.set_appearance_mode( "dark" )
        ctk.set_default_color_theme( "dark-blue" )
        self.width = 1200
        self.height = 700
        self.root = ctk.CTk()
        self.root.title( "Task Monitor" )
        self.root.iconbitmap( os.path.join( os.getcwd(), "Design\schedule.ico" ))
        self.root.geometry( "1200x700+200+80" )
        self.root.resizable( False, False )
        self.path = os.path.join( os.getcwd(), r"Data_File\task_file.xlsx")
        xl = pd.ExcelFile( self.path )
        self.all_sheets = xl.sheet_names
    
    def change( self, can, page) :

        # Switching canvas
        can.destroy()
        page()

    def Imgo( self, file, w, h) :

        # Image processing
        img = Image.open( file )
        pht = ImageTk.PhotoImage( img.resize( (w,h), Image.Resampling.LANCZOS))
        return pht

    def taskAnalysis(self) :

        task_sheet = pd.read_excel( pd.ExcelFile( self.path ), self.all_sheets[1])
        row, col = task_sheet.shape
        task_col = task_sheet.columns
        size = 21

        if ( row<=size ) :
            x, y, z = [],[],[]
            for i in task_sheet[task_col[0]] :
                z.append(i)
            for i in task_sheet[task_col[1]] :
                x.append(i)
            for i in task_sheet[task_col[2]] :
                y.append(i)
            plt.plot( [ i for i in range(row)], x, label = "Assigned Tasks", linewidth = '7', color = 'red')
            plt.plot( [ i for i in range(row)], y, label = "Completed Tasks", linewidth = '7', color = 'green')
            plt.xticks( [ i for i in range(row)], z, rotation = 40)
        
        else :
            diff = row - size
            x, y, z = [],[],[]
            for i in task_sheet[task_col[0]][diff:row] :
                z.append(i)
            for i in task_sheet[task_col[1]][diff:row] :
                x.append(i)
            for i in task_sheet[task_col[2]][diff:row] :
                y.append(i)
            plt.plot( [ i for i in range(size)], x, label = "Assigned Tasks", linewidth = '7', color = 'red')
            plt.plot( [ i for i in range(size)], y, label = "Completed Tasks", linewidth = '7', color = 'green')
            plt.xticks( [ i for i in range(size)], z, rotation = 45)
        plt.xlabel("Dates")
        plt.ylabel("Number of Task")
        plt.title("Task Monitoring Chart")
        plt.legend()
        plt.show()

    def insertTaskAnalysis(self) :

        from datetime import datetime, date
        task_sheet1 = pd.read_excel( pd.ExcelFile( self.path ), self.all_sheets[0])
        task_sheet2 = pd.read_excel( pd.ExcelFile( self.path ), self.all_sheets[1])
        row1, col1 = task_sheet1.shape
        row2, col2 = task_sheet2.shape
        wb = oxl.load_workbook( self.path )
        sheet_xl1 = wb[self.all_sheets[0]]
        sheet_xl2 = wb[self.all_sheets[1]]

        tarik = date.today().strftime("%d/%m/%Y")
        done = ( task_sheet1["Status"] == "Done" )
        done = dict(done.value_counts())
        if True in done.keys() :
            done = done[True]
        else :
            done = 0
        
        if ( sheet_xl2[f"A{row2+1}"].value != tarik ) :

            sheet_xl2[f"A{row2+2}"].value = tarik
            sheet_xl2[f"B{row2+2}"].value = row1 - done
            sheet_xl2[f"C{row2+2}"].value = 0
            sheet_xl2[f"C{row2+1}"].value = done

            count = 2
            task_col = task_sheet1.columns

            for i in range( row1 ) :
                if ( task_sheet1[task_col[2]][i] != "Done" ) :
                    sheet_xl1[f"A{count}"].value = count - 1
                    sheet_xl1[f"B{count}"].value = task_sheet1[task_col[1]][i]
                    sheet_xl1[f"C{count}"].value = task_sheet1[task_col[2]][i]
                    count = count+1
            
            for i in range( count-2, row1 ) :
                sheet_xl1[f"A{i+2}"].value = None
                sheet_xl1[f"B{i+2}"].value = None
                sheet_xl1[f"C{i+2}"].value = None
        
        else :

            sheet_xl2[f"A{row2+1}"].value = tarik
            sheet_xl2[f"B{row2+1}"].value = row1
            sheet_xl2[f"C{row2+1}"].value = done

        wb.save( self.path )

    def insertTask( self, area ) :

        task_sheet = pd.read_excel( pd.ExcelFile( self.path ), self.all_sheets[0] )
        row, col = task_sheet.shape
        task_col = task_sheet.columns

        area.configure( state = "normal")
        for i in range(row) :
            if ( task_sheet[task_col[2]][i] == "Done" ) :
                text = f"✔ {task_sheet[task_col[0]][i]} {task_sheet[task_col[1]][i]} \n"
            else :
                text = f"     {task_sheet[task_col[0]][i]} {task_sheet[task_col[1]][i]} \n"
            area.insert( f"{task_sheet[task_col[0]][i]}.0", text )
        area.configure( state = "disabled")

    def removeTask( self, indx, area, page) :
        
        task_sheet = pd.read_excel( pd.ExcelFile( self.path ), self.all_sheets[0])
        row, col = task_sheet.shape
        
        wb = oxl.load_workbook( self.path )
        sheet_xl = wb[self.all_sheets[0]]
        
        if ( indx != "" ) :

            indx = int(indx)
            if ( indx>0 ) and ( indx<=row ) :

                if ( sheet_xl[f"C{indx+1}"].value != "Done" ) :

                    for i in range( indx+1, row+1 ) :
                        sheet_xl[f"A{i}"].value = sheet_xl[f"A{i+1}"].value - 1
                        sheet_xl[f"B{i}"].value = sheet_xl[f"B{i+1}"].value
                    
                    sheet_xl[f"A{row+1}"].value = None
                    sheet_xl[f"B{row+1}"].value = None

                    try :

                        wb.save( self.path )
                        area.destroy()
                        area = ctk.CTkTextbox( page, 
                                                width = 800, height = 380, 
                                                 text_font = ( "Georgia", 20  ), 
                                                  state = "disabled", corner_radius = 18,
                                                   fg_color = "#6d6e70", bg_color = "#000000",
                                                    border_color = "#3e3e40" , border_width = 8  )
                        area.place( x = 80, y = 230, anchor = "nw")
                        self.insertTask( area )
                        self.insertTaskAnalysis()
                    
                    except :
                        showerror( message = "Close Program related Files", title = "Open File found")

                else :
                    showerror( message = "Marked Task auto Remove Tommorow!", title = "Marked")
            
            else :
                showerror( message = "Invalid Entry!", title = "Invalid")
            
        else :
            showerror( message = "Field Empty!!", title = "Value Not Found")

    def statusTask( self, indx, area, page ) :

        task_sheet = pd.read_excel( pd.ExcelFile( self.path ), self.all_sheets[0])
        row, col = task_sheet.shape
        
        wb = oxl.load_workbook( self.path )
        sheet_xl = wb[self.all_sheets[0]]

        if ( indx != "" ) :

            indx = int(indx)
            if ( indx>0 ) and ( indx<=row ) :

                if ( sheet_xl[f"C{indx+1}"].value != "Done" ) :

                    sheet_xl[f"C{indx+1}"].value = "Done"
                    wb.save( self.path )
                    self.insertTaskAnalysis()
                    area.destroy()
                    area = ctk.CTkTextbox( page, 
                                            width = 800, height = 380, 
                                             text_font = ( "Georgia", 20  ), 
                                              state = "disabled", corner_radius = 18,
                                               fg_color = "#6d6e70", bg_color = "#000000",
                                                border_color = "#3e3e40" , border_width = 8  )
                    area.place( x = 80, y = 230, anchor = "nw")
                    self.insertTask( area )
                
                else :
                    showerror( title = "Invalid", message = "Already Marked" )
        
            else :
                showerror( title = "Invalid!", message = "Invalid Entry" )

        else :
            showerror( message = "Field Empty!!", title = "Value Not Found")

    def updateTask( self, indx, task, area, page) :

        task_sheet = pd.read_excel( pd.ExcelFile( self.path ), self.all_sheets[0])
        row, col = task_sheet.shape
        
        wb = oxl.load_workbook( self.path )
        sheet_xl = wb[self.all_sheets[0]]

        if ( indx != "" ) and ( len(task) > 0 ) :
        
            try :
                indx = int(indx)
                if ( indx == 1 ) :
                    for i in range( row+2, 2, -1 ) :
                        sheet_xl[f"A{i}"] = int(sheet_xl[f"A{i-1}"].value) + 1
                        sheet_xl[f"B{i}"] = sheet_xl[f"B{i-1}"].value
                        sheet_xl[f"C{i}"] = sheet_xl[f"C{i-1}"].value
                    
                    sheet_xl[f"A{2}"] = indx
                    sheet_xl[f"B{2}"] = task
                    sheet_xl[f"C{2}"] = ""
                
                elif ( indx > 1 )  :
                    for i in range( row+2, indx, -1 ) :
                        sheet_xl[f"A{i}"] = int(sheet_xl[f"A{i-1}"].value) + 1
                        sheet_xl[f"B{i}"] = sheet_xl[f"B{i-1}"].value
                        sheet_xl[f"C{i}"] = sheet_xl[f"C{i-1}"].value
                    
                    sheet_xl[f"A{indx+1}"] = indx
                    sheet_xl[f"B{indx+1}"] = task
                    sheet_xl[f"C{indx+1}"] = ""
                
                else :
                    showerror( message = "Invalid Entry!", title = "Invalid")
                
                try :

                    wb.save( self.path )
                    area.destroy()
                    area = ctk.CTkTextbox( page, 
                                            width = 800, height = 380, 
                                             text_font = ( "Georgia", 20  ), 
                                              state = "disabled", corner_radius = 18,
                                               fg_color = "#6d6e70", bg_color = "#000000",
                                                border_color = "#3e3e40" , border_width = 8  )
                    area.place( x = 80, y = 230, anchor = "nw")
                    self.insertTask( area )
                    self.insertTaskAnalysis()
                
                except :
                    showerror( message = "Close Program related Files", title = "Open File found")
            
            except :
                showerror( message = "Insert values of Valid Type", title = "Invalid value found")
        
        else :
            showerror( message = "Field Empty!", title = "Value Not Found")

    def taskMonitoringPage(self) :

        # Defining Structure
        taskMon_page = Canvas( self.root, 
                                width = self.width, height = self.height, 
                                 bg = "black", highlightcolor = "#3c5390", 
                                  borderwidth = 0 )
        taskMon_page.pack( fill = "both", expand = True )

        # Background Image
        back_image = self.Imgo( os.path.join( os.getcwd(), r"Design\secondpage.jpg" ), 1498, 875)
        taskMon_page.create_image( 0, 0, image = back_image , anchor = "nw")
        design_image = self.Imgo( os.path.join( os.getcwd(), r"Design\message.png" ), 70, 75)
        taskMon_page.create_image( 1380, 30, image = design_image , anchor = "nw")

        # Heading
        taskMon_page.create_text( 400, 130, text = "Task Monitoring", 
                                font = ( "Georgia", 42, "bold" ), fill = "#a7a8ac" )

        # Task Index Box
        indx = ctk.CTkEntry( master = taskMon_page, 
                              placeholder_text = "Index", text_font = ( "Georgia", 20 ), 
                               width = 100, height = 30, corner_radius = 14,
                                placeholder_text_color = "#666666", text_color = "#191919", 
                                 fg_color = "#e1f5ff", bg_color = "#000000", 
                                  border_color = "#3e3e40" , border_width = 5 )
        indx_win = taskMon_page.create_window( 100, 200, anchor = "nw", window = indx )
        
        # Task Entry Box
        task = ctk.CTkEntry( master = taskMon_page, 
                              placeholder_text = "Enter Task", text_font = ( "Georgia", 20 ), 
                               width = 550, height = 30, corner_radius = 14,
                                placeholder_text_color = "#666666", text_color = "#191919", 
                                 fg_color = "#e1f5ff", bg_color = "#000000", 
                                  border_color = "#3e3e40" , border_width = 5 )
        task_win = taskMon_page.create_window( 231, 200, anchor = "nw", window = task )

        task_box = ctk.CTkTextbox( taskMon_page, 
                                    width = 800, height = 380, 
                                     text_font = ( "Georgia", 20  ), 
                                      state = "disabled", corner_radius = 18,
                                       fg_color = "#6d6e70", bg_color = "#000000",
                                        border_color = "#3e3e40" , border_width = 8  )
        task_box.place( x = 80, y = 230, anchor = "nw")

        self.insertTaskAnalysis()
        self.insertTask( task_box )

        task.bind('<Return>', lambda event = None : self.updateTask( indx.get(), task.get(), task_box, taskMon_page ) )

        # Insert Button
        insert_bt = ctk.CTkButton( master = taskMon_page, 
                                   text = "Insert", text_font = ( "Georgia", 20  ), 
                                    width = 100, height = 40, corner_radius = 18,
                                     bg_color = "#000000", fg_color = "red", 
                                      hover_color = "#ff5359", border_width = 0, 
                                       command = lambda : self.updateTask( indx.get(), task.get(), task_box, taskMon_page ) )
        insert_bt_win = taskMon_page.create_window( 930, 200, anchor = "nw", window = insert_bt )

        # Remove Task
        taskMon_page.create_text( 1360, 318, text = "Remove Task", 
                                font = ( "Tahoma", 18, "italic", "underline" ), fill = "white" )

        # Task Remove Box
        remove = ctk.CTkEntry( master = taskMon_page, 
                                placeholder_text = "Index", text_font = ( "Georgia", 20  ), 
                                 width = 100, height = 30, corner_radius = 14,
                                  placeholder_text_color = "#666666", text_color = "#191919", 
                                   fg_color = "#e1f5ff", bg_color = "#000000", 
                                    border_color = "#3b43c8" , border_width = 5 )
        remove_win = taskMon_page.create_window( 1330, 348, anchor = "nw", window = remove )

        remove.bind('<Return>', lambda event = None : self.removeTask( remove.get(), task_box, taskMon_page ) )

        # Remove Button
        remove_bt = ctk.CTkButton( master = taskMon_page, 
                                    text = "Remove", text_font = ( "Georgia", 20  ), 
                                     width = 80, height = 40, corner_radius = 18,
                                      bg_color = "#000000", fg_color = "red", 
                                       hover_color = "#ff5359", border_width = 0, 
                                        command = lambda : self.removeTask( remove.get(), task_box, taskMon_page ) )
        remove_bt_win = taskMon_page.create_window( 1290, 410, anchor = "nw", window = remove_bt )

        # Mark Done Task
        taskMon_page.create_text( 1350, 528, text = "Mark Done Task", 
                                font = ( "Tahoma", 18, "italic", "underline" ), fill = "white" )

        # Task Done Box
        tkdone = ctk.CTkEntry( master = taskMon_page, 
                                placeholder_text = "Index", text_font = ( "Georgia", 20 ), 
                                 width = 100, height = 30, corner_radius = 14,
                                  placeholder_text_color = "#666666", text_color = "#191919", 
                                   fg_color = "#e1f5ff", bg_color = "#000000", 
                                    border_color = "#3b43c8" , border_width = 5 )
        tkdone_win = taskMon_page.create_window( 1330, 558, anchor = "nw", window = tkdone )

        tkdone.bind('<Return>', lambda event = None : self.statusTask( tkdone.get(), task_box, taskMon_page ) )
    
        # Done Button
        tkdone_bt = ctk.CTkButton( master = taskMon_page, 
                                   text = "Mark Done", text_font = ( "Georgia", 20 ), 
                                    width = 70, height = 40, corner_radius = 18,
                                     bg_color = "#000000", fg_color = "red", 
                                      hover_color = "#ff5359", border_width = 0, 
                                       command = lambda : self.statusTask( tkdone.get(), task_box, taskMon_page ) )
        tkdone_bt_win = taskMon_page.create_window( 1240, 620, anchor = "nw", window = tkdone_bt )

        # Analysis Button
        analysis_bt = ctk.CTkButton( master = taskMon_page, 
                                      text = "Analysis", text_font = ( "Georgia", 20 ), 
                                       width = 120, height = 40, corner_radius = 18,
                                        bg_color = "#000000", fg_color = "red", 
                                         hover_color = "#ff5359", border_width = 0, 
                                          command = lambda : self.taskAnalysis() )
        analysis_bt_win = taskMon_page.create_window( 650, 790, anchor = "nw", window = analysis_bt )

        # Return Button
        ret_img = self.Imgo( os.path.join( os.getcwd(), "Design\logout.png" ), 35, 35 )
        back_bt = ctk.CTkButton( master = taskMon_page, 
                                  image = ret_img, text = None, 
                                   width = 45, height = 45, corner_radius = 23, 
                                    bg_color = "#000000", fg_color = "red", 
                                     hover_color = "#ff5359", border_width = 0,  
                                      command = lambda : self.change( taskMon_page, self.firstPage ))
        back_bt_win = taskMon_page.create_window( 30, 20, anchor = "nw", window = back_bt )

        self.root.mainloop()

    def firstPage(self) :

        # Defining Structure
        first_page = Canvas( self.root, 
                              width = self.width, height = self.height, 
                               bg = "black", highlightcolor = "#3c5390", 
                                borderwidth = 0 )
        first_page.pack( fill = "both", expand = True )

        # Background Image
        back_image = self.Imgo( os.path.join( os.getcwd(), r"Design\firstpage.jpg" ), 1498, 875)
        first_page.create_image( 0, 0, image = back_image , anchor = "nw")

        # Heading
        first_page.create_text( 300, 170, text = "Task", 
                                font = ( "Georgia", 60, "bold" ), fill = "#a7a8ac" )
        first_page.create_text( 390, 280, text = "Monitor", 
                                font = ( "Georgia", 60, "bold" ), fill = "#e3e3e3" )#222222

        # Next Page Button
        next_bt = ctk.CTkButton( master = first_page,
                                  text = "Let's  Go  ->", text_font = ( "Georgia", 23 ),
                                   width = 120, height = 50, corner_radius = 14,
                                    bg_color = "#0d0d0d", fg_color = "red",
                                     hover_color = "#ff5359", border_width = 0,
                                      text_color = "white",
                                       command = lambda : self.change( first_page, self.taskMonitoringPage ) )
        next_bt_win = first_page.create_window( 190, 640, anchor = "nw", window = next_bt )

        self.root.mainloop()

if __name__ == "__main__" :

    task_class = TaskMonitor()
    task_class.firstPage()
